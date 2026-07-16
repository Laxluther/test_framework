from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path
import json

# --- Styles ---
GREEN      = PatternFill("solid", fgColor="8FD14F")
YELLOW     = PatternFill("solid", fgColor="FFD580")
RED        = PatternFill("solid", fgColor="FC8A8C")
BLUE_HDR   = PatternFill("solid", fgColor="1F4E79")
PURPLE_HDR = PatternFill("solid", fgColor="4E1F79")
GREY_HDR   = PatternFill("solid", fgColor="D9D9D9")
ALT_ROW    = PatternFill("solid", fgColor="EBF3FB")
WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BLACK_BOLD = Font(bold=True, name="Arial", size=10)
CELL_FONT  = Font(name="Arial", size=10)
WRAP       = Alignment(wrap_text=True, vertical="top")
CENTRE     = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN       = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

GRADES_HEADERS = ["No.", "Application", "Expected Grades", "Suggested Grades", "Matched Grades",
                  "Total Suggested", "Total Expected", "Total Matched", "Grade Pass", "Flow Completed", "Error"]
GRADES_WIDTHS = [6, 22, 38, 38, 30, 14, 13, 13, 11, 14, 30]

ASSUMPTION_HEADERS = ["No.", "Application", "Pass/Fail", "Score /10", "Matched", "Total Exp",
                      "Expected CTQs", "Agent Assumption Output", "Matched CTQs", "Unmatched CTQs",
                      "Extra CTQs", "Match Details (Evidence)", "Unmatched Reasons", "Method", "Reasoning"]
ASSUMPTION_WIDTHS = [5, 28, 10, 10, 9, 9, 38, 55, 38, 35, 38, 55, 45, 12, 55]

OVERVIEW_SUMMARY_HEADERS = ["PASS", "FAIL", "Score", ">80%", ">70%", ">60%"]

TIMING_HEADERS = ["No.", "Application", "Turn", "User Input", "Response Time (ms)", "Agent / Tool", "Type", "Call Duration (ms)", "Succeeded", "Tool Input", "Tool Output"]
TIMING_WIDTHS = [6, 22, 7, 45, 16, 30, 10, 16, 11, 45, 45]

def _pass_fill(passed: bool | None) -> PatternFill:
    if passed is None: return PatternFill(fill_type=None)
    return GREEN if passed else RED

def _score_fill(score: float | None) -> PatternFill:
    if score is None: return PatternFill(fill_type=None)
    if score >= 7.0: return GREEN
    if score >= 5.0: return YELLOW
    return RED

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

def _hdr(ws, headers, row, fill):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = WHITE_BOLD
        cell.fill = fill
        cell.alignment = CENTRE
        cell.border = THIN

def _fmt_list(items):
    if not items: return ""
    return "\n".join(f"• {x}" for x in items)

def _fmt_match_details(details):
    if not details: return ""
    return "\n".join(f"• {d.get('expected', '')}: {d.get('actualEvidence', '')}" for d in details)

def _fmt_unmatched_reasons(reasons):
    if not reasons: return ""
    return "\n".join(f"• {r.get('expected', '')}: {r.get('reason', '')}" for r in reasons)

def _grade_status(res: dict) -> str:
    eval_data = res.get("gradeEvaluation") or res.get("evaluation") or {}
    passed = eval_data.get("passed")
    return "PASS" if passed else ("FAIL" if passed is False else "N/A")

def build_grades_sheet(ws, results: list[dict]):
    _set_col_widths(ws, GRADES_WIDTHS)
    _hdr(ws, GRADES_HEADERS, 1, PURPLE_HDR)

    for r_idx, res in enumerate(results, 2):
        eval_data = res.get("gradeEvaluation", {})
        if not eval_data:
            eval_data = res.get("evaluation", {}) # Fallback for old format

        row_data = [
            res.get("conversationNo", ""),
            res.get("application", ""),
            _fmt_list(res.get("expectedGrades", [])),
            _fmt_list([g.get("gradeName", str(g)) if isinstance(g, dict) else str(g) for g in res.get("suggestedGrades", [])]),
            _fmt_list(eval_data.get("matchedSuggested", [])),
            eval_data.get("totalSuggested", 0),
            eval_data.get("totalExpected", 0),
            eval_data.get("totalMatched", 0),
            "PASS" if eval_data.get("passed") else ("FAIL" if eval_data.get("passed") is False else "N/A"),
            "Yes" if res.get("flowCompleted") else "No",
            res.get("error", "")
        ]

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c_idx % 2 == 0: cell.fill = ALT_ROW

            if GRADES_HEADERS[c_idx-1] == "Grade Pass":
                cell.fill = _pass_fill(eval_data.get("passed"))
            if GRADES_HEADERS[c_idx-1] == "Flow Completed":
                cell.fill = _pass_fill(res.get("flowCompleted"))

def build_assumption_sheet(ws, results: list[dict]):
    _set_col_widths(ws, ASSUMPTION_WIDTHS)
    _hdr(ws, ASSUMPTION_HEADERS, 1, BLUE_HDR)

    for r_idx, res in enumerate(results, 2):
        eval_data = res.get("assumptionEvaluation", {})
        if not eval_data:
            continue

        row_data = [
            res.get("conversationNo", ""),
            res.get("application", ""),
            "PASS" if eval_data.get("passed") else ("FAIL" if eval_data.get("passed") is False else "N/A"),
            eval_data.get("overallScore", ""),
            len(eval_data.get("matchedCTQs", [])),
            len(eval_data.get("matchedCTQs", [])) + len(eval_data.get("unmatchedCTQs", [])),
            _fmt_list([c.get("expected", "") for c in eval_data.get("matchedCTQs", [])] + [c.get("expected", "") for c in eval_data.get("unmatchedCTQs", [])]),
            res.get("agentAssumptionOutput", ""),
            _fmt_list([c.get("expected", "") for c in eval_data.get("matchedCTQs", [])]),
            _fmt_list([c.get("expected", "") for c in eval_data.get("unmatchedCTQs", [])]),
            _fmt_list(eval_data.get("extraCTQs", [])),
            _fmt_match_details(eval_data.get("matchedCTQs", [])),
            _fmt_unmatched_reasons(eval_data.get("unmatchedCTQs", [])),
            eval_data.get("method", ""),
            eval_data.get("reasoning", "")
        ]

        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c_idx % 2 == 0: cell.fill = ALT_ROW

            if ASSUMPTION_HEADERS[c_idx-1] == "Pass/Fail":
                cell.fill = _pass_fill(eval_data.get("passed"))
            if ASSUMPTION_HEADERS[c_idx-1] == "Score /10":
                cell.fill = _score_fill(eval_data.get("overallScore"))

def _fmt_ms(ms):
    return round(ms, 1) if ms is not None else ""

def build_timing_sheet(ws, results: list[dict]):
    """Per-round timing breakdown: for every conversation, every turn's input/output
    response time, and (when MLflow trace data was captured) which agents/tools ran
    inside that turn and how long each took. Ends with round-level overall/avg summary."""
    _set_col_widths(ws, TIMING_WIDTHS)
    _hdr(ws, TIMING_HEADERS, 1, GREY_HDR)

    def _write_row(r_idx, row_data, succeeded=None):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = THIN
            if c_idx % 2 == 0:
                cell.fill = ALT_ROW
            if TIMING_HEADERS[c_idx - 1] == "Succeeded" and succeeded is not None:
                cell.fill = _pass_fill(succeeded)

    r_idx = 2
    round_total_ms = 0.0
    round_total_count = 0
    turn_latencies = []

    for res in results:
        conv_no = res.get("conversationNo", "")
        app_name = res.get("application", "")
        turn_traces = res.get("turnTraces") or []

        conv_total = (res.get("timing") or {}).get("totalDurationMs")
        if conv_total is not None:
            round_total_ms += conv_total
            round_total_count += 1

        if not turn_traces:
            _write_row(r_idx, [conv_no, app_name, "", "(no turn trace data captured for this run)", "", "", "", "", "", "", ""])
            r_idx += 1
            continue

        for turn in turn_traces:
            response_ms = turn.get("responseTimeMs")
            if response_ms is not None:
                turn_latencies.append(response_ms)
            agent_calls = turn.get("agentCalls") or []
            if not agent_calls:
                _write_row(r_idx, [conv_no, app_name, turn.get("turnNo", ""), turn.get("userInput", ""), _fmt_ms(response_ms), "", "", "", "", "", ""])
                r_idx += 1
            else:
                for call in agent_calls:
                    succeeded = call.get("succeeded")
                    _write_row(r_idx, [
                        conv_no, app_name, turn.get("turnNo", ""), turn.get("userInput", ""), _fmt_ms(response_ms),
                        call.get("name", ""), call.get("type", ""), _fmt_ms(call.get("durationMs")),
                        ("Yes" if succeeded else "No") if succeeded is not None else "",
                        call.get("inputs", ""), call.get("outputs", ""),
                    ], succeeded=succeeded)
                    r_idx += 1

    r_idx += 1
    avg_turn_ms = sum(turn_latencies) / len(turn_latencies) if turn_latencies else None
    for label, val in [
        ("Round Overall Time (sum of conversation durations)", _fmt_ms(round_total_ms) if round_total_count else ""),
        ("Avg Turn Response Time (this round)", _fmt_ms(avg_turn_ms) if avg_turn_ms is not None else ""),
    ]:
        label_cell = ws.cell(row=r_idx, column=1, value=label)
        label_cell.font = WHITE_BOLD
        label_cell.fill = BLUE_HDR
        label_cell.border = THIN
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=4)
        val_cell = ws.cell(row=r_idx, column=5, value=val)
        val_cell.font = BLACK_BOLD
        val_cell.border = THIN
        r_idx += 1

    ws.freeze_panes = "A2"

def build_overview_sheet(ws, results_by_round: dict[int, list[dict]]):
    """Cross-round PASS/FAIL matrix: one row per conversation, one column per
    round, with row-wise and column-wise summaries. Mirrors the 'overview'
    sheet the team already relies on (No. | Application | round1..roundN |
    PASS | FAIL | Score | >80% | >70% | >60%, plus PASS/FAIL/Score totals)."""
    round_nums = sorted(results_by_round.keys())

    conv_apps = {}
    for rnd in round_nums:
        for res in results_by_round[rnd]:
            conv_no = res.get("conversationNo")
            if conv_no is None:
                continue
            conv_apps.setdefault(conv_no, res.get("application", ""))
    conv_nos = sorted(conv_apps.keys())

    status_lookup = {}
    for rnd in round_nums:
        for res in results_by_round[rnd]:
            conv_no = res.get("conversationNo")
            if conv_no is None:
                continue
            status_lookup[(conv_no, rnd)] = _grade_status(res)

    headers = ["No.", "Application"] + [f"round{r}" for r in round_nums] + OVERVIEW_SUMMARY_HEADERS
    widths = [6, 24] + [10] * len(round_nums) + [8, 8, 10, 9, 9, 9]
    _set_col_widths(ws, widths)
    _hdr(ws, headers, 1, BLUE_HDR)

    round_tally = {r: {"PASS": 0, "FAIL": 0} for r in round_nums}
    threshold_tally = {">80%": {"Yes": 0, "No": 0}, ">70%": {"Yes": 0, "No": 0}, ">60%": {"Yes": 0, "No": 0}}
    per_app_scores = []

    for r_idx, conv_no in enumerate(conv_nos, 2):
        row_vals = [conv_no, conv_apps[conv_no]]
        pass_count = 0
        fail_count = 0
        for rnd in round_nums:
            status = status_lookup.get((conv_no, rnd), "N/A")
            row_vals.append(status)
            if status == "PASS":
                pass_count += 1
                round_tally[rnd]["PASS"] += 1
            elif status == "FAIL":
                fail_count += 1
                round_tally[rnd]["FAIL"] += 1

        graded_total = pass_count + fail_count
        score = (pass_count / graded_total * 100) if graded_total > 0 else 0.0
        per_app_scores.append(score)
        gt80 = "Yes" if score >= 80 else "No"
        gt70 = "Yes" if score >= 70 else "No"
        gt60 = "Yes" if score >= 60 else "No"
        threshold_tally[">80%"][gt80] += 1
        threshold_tally[">70%"][gt70] += 1
        threshold_tally[">60%"][gt60] += 1

        row_vals += [pass_count, fail_count, f"{score:.1f}%", gt80, gt70, gt60]

        for c_idx, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = CENTRE
            cell.border = THIN
            if val in ("PASS", "Yes"):
                cell.fill = GREEN
            elif val in ("FAIL", "No"):
                cell.fill = RED

    n_round_cols = len(round_nums)
    pass_col = 3 + n_round_cols
    fail_col = pass_col + 1
    score_col = fail_col + 1
    summary_row_start = len(conv_nos) + 3
    overall_score = sum(per_app_scores) / len(per_app_scores) if per_app_scores else 0.0

    for i, label in enumerate(["PASS", "FAIL", "Score"]):
        r_idx = summary_row_start + i
        label_cell = ws.cell(row=r_idx, column=2, value=label)
        label_cell.font = WHITE_BOLD
        label_cell.fill = BLUE_HDR
        label_cell.border = THIN
        ws.cell(row=r_idx, column=1).border = THIN

        for j, rnd in enumerate(round_nums):
            p, f = round_tally[rnd]["PASS"], round_tally[rnd]["FAIL"]
            if label == "PASS": val = p
            elif label == "FAIL": val = f
            else:
                total = p + f
                val = f"{(p / total * 100):.1f}%" if total > 0 else "0.0%"
            cell = ws.cell(row=r_idx, column=3 + j, value=val)
            cell.font = CELL_FONT
            cell.alignment = CENTRE
            cell.border = THIN

        ws.cell(row=r_idx, column=pass_col).border = THIN
        ws.cell(row=r_idx, column=fail_col).border = THIN

        score_cell = ws.cell(row=r_idx, column=score_col)
        score_cell.border = THIN
        score_cell.alignment = CENTRE
        if label == "Score":
            score_cell.value = f"{overall_score:.1f}%"
            score_cell.font = BLACK_BOLD

        for k, thresh in enumerate([">80%", ">70%", ">60%"]):
            cell = ws.cell(row=r_idx, column=score_col + 1 + k)
            cell.border = THIN
            cell.alignment = CENTRE
            cell.font = CELL_FONT
            if label == "PASS":
                cell.value = threshold_tally[thresh]["Yes"]
            elif label == "FAIL":
                cell.value = threshold_tally[thresh]["No"]
            else:
                yes, no = threshold_tally[thresh]["Yes"], threshold_tally[thresh]["No"]
                total = yes + no
                cell.value = f"{(yes / total * 100):.1f}%" if total > 0 else "0.0%"
                cell.font = BLACK_BOLD

    ws.freeze_panes = "C2"

def generate_report_from_results_by_round(results_by_round: dict[int, list[dict]], output_path: str | Path) -> Path:
    """Build the consolidated workbook (overview + round{N}/assumption_round{N}
    sheets) directly from an in-memory {round_no: [result_dict, ...]} mapping.
    Shared by the file-based generate_report() and by DB-driven regeneration
    (a session whose results/ folder was cleaned up can still be exported)."""
    output_path = Path(output_path)
    wb = Workbook()
    wb.remove(wb.active)

    if results_by_round:
        ws_overview = wb.create_sheet(title="overview")
        build_overview_sheet(ws_overview, results_by_round)

    for round_no in sorted(results_by_round.keys()):
        results = results_by_round[round_no]

        ws_grades = wb.create_sheet(title=f"round{round_no}")
        build_grades_sheet(ws_grades, results)

        has_assumptions = any(r.get("assumptionEvaluation") for r in results)
        if has_assumptions:
            ws_assumptions = wb.create_sheet(title=f"assumption_round{round_no}")
            build_assumption_sheet(ws_assumptions, results)

        ws_timing = wb.create_sheet(title=f"timing_round{round_no}")
        build_timing_sheet(ws_timing, results)

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    wb.save(output_path)
    return output_path

def generate_report(results_folder: str | Path, output_path: str | Path) -> Path:
    results_folder = Path(results_folder)

    round_dirs = sorted(
        [d for d in results_folder.iterdir() if d.is_dir() and d.name.startswith("round")],
        key=lambda d: int(d.name.replace("round", "") or 0)
    )
    if not round_dirs:
        round_dirs = [results_folder]

    results_by_round: dict[int, list[dict]] = {}
    for round_dir in round_dirs:
        round_name = round_dir.name if round_dir.name.startswith("round") else "round1"
        round_no = int(round_name.replace("round", "") or 1)

        results = []
        for file_path in round_dir.glob("result_*.json"):
            with open(file_path, "r", encoding="utf-8") as f:
                results.append(json.load(f))

        if not results:
            continue

        results.sort(key=lambda x: x.get("conversationNo", 999))
        results_by_round[round_no] = results

    return generate_report_from_results_by_round(results_by_round, output_path)
